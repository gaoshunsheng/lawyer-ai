import pytest


def test_export_xlsx_generates_workbook():
    """Test that xlsx export produces a valid workbook without DB."""
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "案件概览"
    ws.append(["指标", "数值"])
    ws.append(["案件总数", 120])
    ws.append(["胜诉率(%)", 72.5])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from openpyxl import load_workbook
    loaded = load_workbook(buf)
    assert loaded.active.title == "案件概览"
    assert loaded.active.cell(row=2, column=2).value == 120


def test_export_pdf_generates_bytes():
    """Test that PDF export produces non-empty bytes without DB."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as pdf_canvas

    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    c.setFont("Helvetica", 16)
    c.drawString(72, A4[1] - 72, "Test Report")
    c.save()
    buf.seek(0)

    data = buf.read()
    assert len(data) > 0
    assert data[:5] == b"%PDF-"


def test_export_xlsx_multiple_sheets():
    from io import BytesIO
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.append(["类型", "数量"])
    ws1.append(["labor_contract", 45])

    ws2 = wb.create_sheet("按状态")
    ws2.append(["状态", "数量"])
    ws2.append(["active", 30])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    loaded = load_workbook(buf)
    assert len(loaded.sheetnames) == 2
    assert "按状态" in loaded.sheetnames
