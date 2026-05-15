from __future__ import annotations

from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.dependencies import get_current_user, require_role
from app.models import User
from app.services import report_service

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/case-overview")
async def case_overview(
    days: int = Query(90, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    start = datetime.now(timezone.utc) - timedelta(days=days)
    return await report_service.get_case_overview(db, current_user.tenant_id, start)


@router.get("/trends")
async def trends(
    granularity: str = Query("month", pattern="^(month|quarter)$"),
    days: int = Query(180, ge=30, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await report_service.get_trends(db, current_user.tenant_id, granularity, days)


@router.get("/lawyer-performance")
async def lawyer_performance(
    period_days: int = Query(90, ge=30, le=365),
    current_user: User = Depends(require_role("platform_admin", "tenant_admin")),
    db: AsyncSession = Depends(get_db),
):
    return await report_service.get_lawyer_performance(db, current_user.tenant_id, period_days)


@router.get("/export")
async def export_report(
    format: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    report_type: str = Query("case_overview"),
    days: int = Query(90, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    start = datetime.now(timezone.utc) - timedelta(days=days)
    data = await report_service.get_case_overview(db, current_user.tenant_id, start)

    if format == "xlsx":
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "案件概览"
        ws.append(["指标", "数值"])
        ws.append(["案件总数", data["total"]])
        ws.append(["平均周期(天)", data["avg_duration_days"]])
        ws.append(["胜诉率(%)", data["win_rate"]])
        ws.append(["总标的额", data["total_claim_amount"]])

        if data["by_type"]:
            ws2 = wb.create_sheet("按类型统计")
            ws2.append(["案件类型", "数量", "标的额"])
            for row in data["by_type"]:
                ws2.append([row["type"], row["count"], row["amount"]])

        if data["by_status"]:
            ws3 = wb.create_sheet("按状态统计")
            ws3.append(["案件状态", "数量"])
            for row in data["by_status"]:
                ws3.append([row["status"], row["count"]])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=report.xlsx"},
        )
    else:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfgen import canvas as pdf_canvas

        buf = BytesIO()
        c = pdf_canvas.Canvas(buf, pagesize=A4)
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        c.setFont("STSong-Light", 16)
        c.drawString(72, A4[1] - 72, "案件概览报告")
        c.setFont("STSong-Light", 11)
        y = A4[1] - 110
        c.drawString(72, y, f"案件总数: {data['total']}")
        y -= 20
        c.drawString(72, y, f"平均周期(天): {data['avg_duration_days']}")
        y -= 20
        c.drawString(72, y, f"胜诉率(%): {data['win_rate']}")
        y -= 20
        c.drawString(72, y, f"总标的额: {data['total_claim_amount']}")
        y -= 30
        c.drawString(72, y, "按类型统计:")
        y -= 18
        for row in data["by_type"]:
            c.drawString(90, y, f"{row['type']}: {row['count']}件, {row['amount']}")
            y -= 16
        y -= 10
        c.drawString(72, y, "按状态统计:")
        y -= 18
        for row in data["by_status"]:
            c.drawString(90, y, f"{row['status']}: {row['count']}")
            y -= 16
        c.save()
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=report.pdf"},
        )
